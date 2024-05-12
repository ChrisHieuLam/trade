#DATE: 09/05/2024
#VERSION: 1.0

# Note:
# Import data from tv to SQL server 
import json
import re
import shutil
import pandas as pd
from sqlalchemy.engine import URL
from sqlalchemy.engine import create_engine
from sqlalchemy import text
import pyodbc 
from sqlalchemy import insert
from sqlalchemy import MetaData, Table
import sqlalchemy as db
from sqlalchemy import select
import sys
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
import time
import matplotlib.pyplot as plt
import base64



SERVER_NAME = 'lamfa.webhop.net'
DATABASE_NAME = 'dbDataWhale'
TABLE_NAME = 'tbData'
USER_NAME = 'appcoder'
PASS_WORD = 'Team_500'

conn_str = (
    r'DRIVER={ODBC Driver 17 for SQL Server};'
    r'SERVER=lamfa.webhop.net;'
    r'DATABASE=dbDataWhale;'
    r'UID=appcoder;'
    r'PWD=Team_500;'
)

connection_url = URL.create('mssql+pyodbc', query={'odbc_connect': conn_str})
engine = create_engine(connection_url, module=pyodbc)
connection = engine.connect()

def to_excel_float(date_string):
    if date_string is None:
        return None
    date_obj = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
    base_date = datetime(1899, 12, 30)
    days_since_base = date_obj.toordinal() - base_date.toordinal()
    fraction_of_day = (date_obj - datetime(date_obj.year, date_obj.month, date_obj.day)).total_seconds() / (24 * 3600)
    return days_since_base + fraction_of_day

def format_number_5(number):
    num_str = str(number)
    start, end = num_str.split(".")
    while len(start) < 3:
        start = "0" + start
    while len(end) < 3:
        end = end + "0"
    return start[:3] + end[:2]
def format_number_4(number):
    num_str = str(number)
    start, end = num_str.split(".")
    while len(start) < 2:
        start = "0" + start
    while len(end) < 3:
        end = end + "0"
    return start[:2] + end[:2]
def graph_create(i, symbol):
    trade_date = []
    trade_price = []
    with engine.connect() as connection:
        query = db.select(tbSummary).where(tbSummary.c.StrategyID == i)
        summary = connection.execute(query)
        for row in summary:
            profit_fac = row[12]
            per_pro = row[20]
            avg_trade = row[21]
            max_drawdown = row[8]
            max_runup = row[7]
            net_profit = row[4]
            total_closed_trades = row[13]
            profit_factor = format_number_5(row[12])
            percent_profitable = format_number_4(row[20])
        connection.close()
    with engine.connect() as connection:
        query = db.select(tbSummary.c.PropertiesID).where(tbSummary.c.StrategyID == i)
        summary_prop = connection.execute(query)
        for row in summary_prop:
            properties_ID_file = row[0]
        connection.close()
    with engine.connect() as connection:
        query = db.select(tbStrategy.c.StrategyCode).where(tbStrategy.c.StrategyID == i)
        strat_code = connection.execute(query)
        for row in strat_code:
            strategy_code_file = row[0].strip()
        connection.close()
    with engine.connect() as connection:
        query = db.select(tbProperties).where(tbProperties.c.PropertiesID == properties_ID_file)
        backtest = connection.execute(query)
        for row in backtest:
            timeframe = row[5]
            trading_range = row[2]
            back_testing = row[3]
            strat_inputs = row[-2]
            strat_prop = row[-1]
            start_str, end_str = row[3].split(" — ")
            # Convert the string representations to datetime objects
            try:
                start_date = datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
                end_date = datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")
            except Exception:
                start_date = datetime.strptime(start_str, "%Y-%m-%d %H:%M")
                end_date = datetime.strptime(end_str, "%Y-%m-%d %H:%M")
            # Format start and end dates to YYYYMMDD format
            formatted_start_date = start_date.strftime("%Y%m%d")
            formatted_end_date = end_date.strftime("%Y%m%d")
            # Concatenate formatted dates with an underscore
            backtestingrange = f"{formatted_start_date}_{formatted_end_date}"
        connection.close()
    with engine.connect() as connection:
        query = select(tbTradeList).where(tbTradeList.c.StrategyID == i)
        trade_data = connection.execute(query)
        for row in trade_data:
            trade_date.append(row[5].strftime("%Y-%m-%d %H:%M:%S"))  # Convert datetime to timestamp
            if float(row[6]) != 0:
                trade_price.append(float(row[6]))  # Convert to float
        excel_float_dates = [to_excel_float(date_str) for date_str in trade_date if datetime.strptime(date_str,"%Y-%m-%d %H:%M:%S") > datetime.strptime("2000-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")]
        datetime_dates = [datetime(1899, 12, 30) + timedelta(days=float_date) for float_date in excel_float_dates if float_date]
        plt.figure(figsize=(20, 6))  
        plt.scatter(datetime_dates, trade_price, c=excel_float_dates, cmap='viridis', edgecolor='black', alpha=0.5)
        plt.xlabel('Date')
        plt.ylabel('Price USD')
        plt.title('List of Trades')
        plt.grid(True)
        plt.switch_backend('agg')
        path = os.path.join("/home/appcoder/graph_saved/", f"{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}")
        # Create the new folder
        os.makedirs(path, exist_ok=True)
        plt.savefig(os.path.join(path,f'{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}_price_usd.png'))
        plt.close()
        connection.close()
    with engine.connect() as connection:
        query = select(tbTradeList).where(tbTradeList.c.StrategyID == i)
        trade_data = connection.execute(query)
        profit_usd = []
        for row in trade_data:
            profit_usd.append(float(row[8]))
        profit_usd.pop(0)
        plt.figure(figsize=(20, 6))  
        plt.scatter(datetime_dates, profit_usd, c=excel_float_dates, cmap='viridis', edgecolor='black', alpha=0.5)
        plt.xlabel('Date')
        plt.ylabel('Profit USD')
        plt.title('List of Trades')
        plt.grid(True)
        plt.switch_backend('agg')
        plt.savefig(os.path.join(path,f'{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}_profit_usd.png'))
        plt.close()
        filename = f"{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}.html"
        connection.close()
    with open(os.path.join(path,f'{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}_profit_usd.png'), "rb") as img_file:
        profit_data_pic = img_file.read()
        profit_base64 = base64.b64encode(profit_data_pic).decode("utf-8")
    with open(os.path.join(path,f'{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}_price_usd.png'), "rb") as img_file:
        price_data_pic = img_file.read()
        price_base64 = base64.b64encode(price_data_pic).decode("utf-8")

    # HTML code with embedded image
    html_code = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{symbol}_{strategy_code_file}_{profit_factor}_{percent_profitable}_{backtestingrange}</title>
    </head>
    <body>
    <h1>Infomation</h1>
        Symbol: {symbol}
        <br> <br>  
        Strategy: {strategy_code_file}
        <br> <br> 
        Timeframe: {timeframe}
        <br> 
        Profit Factor: {profit_fac}
        <br> <br> 
        Percent Profitable: {per_pro}
        <br> <br> 
        Avg Trade: {avg_trade}
        <br> <br> 
        Max Drawdown: {max_drawdown}
        <br> <br> 
        Max RunUp: {max_runup}
        <br> <br> 
        Net Profit: {net_profit}
        <br> <br> 
        Total Closed Trades: {total_closed_trades}
        <br> <br> 
        Trading Range: {trading_range}
        <br> <br> 
        Backtesting Range: {back_testing}
        <br> <br> 
        Strategy Inputs: {strat_inputs}
        <br> <br> 
        Strategy Properties: {strat_prop}
        <br> <br> 
    <h1>Graph</h1>
    <img src="data:image/png;base64,{price_base64}" alt="Graph of Price/Date">
    <img src="data:image/png;base64,{profit_base64}" alt="Graph of Price/Date">
    </body>
    </html>
    """
    
    file_path = os.path.join(path, filename)
    # Save HTML to a file
    with open(file_path, "w") as html_file:
        html_file.write(html_code)
while True:
    try:
        excel_path = '/home/appcoder/excels-tv'
        excel_files = [file for file in os.listdir(excel_path) if file.endswith('.xlsx')]
        while not any(os.listdir(excel_path)):
            print("Waiting for file")
            time.sleep(2)
        for workbook in excel_files:
            workbook_path = os.path.join(excel_path, workbook)
            work_book = load_workbook(workbook_path)

            properties_data = {}
            summary_data = {} 
            for sheet_name in work_book.sheetnames:
                sheet = work_book[sheet_name]
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    if "List" in sheet_name:
                        list_data = pd.read_excel(workbook_path, sheet_name = sheet_name).to_dict(orient = "records") 
                    if "Performance" in sheet_name:
                        summary_data[row[0]] = row[1]
                    if "Properties" in sheet_name:
                        properties_data[row[0]] = row[1]
            del properties_data['Title']
            del summary_data['Unnamed: 0']
            for i in list_data:
                for key,value in i.items():
                    if isinstance(value, float) and pd.isna(value):
                        i[key] = 0



            #defining tables in SQLAlchemy

            metadata = db.MetaData() #extracting the metadata
            tbProperties= db.Table('tbProperties', metadata, 
            autoload_with=engine)
            tbSymbol= db.Table('tbSymbol', metadata, 
            autoload_with=engine)
            tbStrategy= db.Table('tbStrategy', metadata, 
            autoload_with=engine)
            tbSummary= db.Table('tbSummary', metadata, 
            autoload_with=engine)
            tbTradeList = db.Table('tbTradeList', metadata, 
            autoload_with=engine)
            metadata.reflect(bind=engine)

            #check if Symbol is inserted or not
            #insert if not
            with engine.connect() as connection:
                select_query = db.select(tbSymbol.c.SymbolCode)
                return_Codes = connection.execute(select_query).fetchall()
                codes = []
                for code in return_Codes:
                    codes.append(code[0])
                symbolname = properties_data['Symbol']
                while len(symbolname) != 20:
                    symbolname += ' '
                if symbolname not in codes:
                    put_in = symbolname.split(":")
                    ins = tbSymbol.insert().values(**{
                    "SymbolCode": properties_data['Symbol'],
                    "SymbolName": put_in[1]
                    }
                    )
                    connection.execute(ins)
                    connection.commit()
                connection.close()

            # insert SymbolID to Properties
            with engine.connect() as connection:
                select_columns = db.select(tbSymbol.c.SymbolID)
                condition = tbSymbol.c.SymbolCode == properties_data['Symbol']
                query = db.select(tbSymbol.c.SymbolID).where(condition)
                output = connection.execute(query)
                #Get Symbol ID WHERE Symbol == properties_data['Symbol']
                inputs = None
                for out in output:
                    inputs = out[0]
                #get start end dates
                dates = properties_data['Trading range'].split("—")
                try:
                    start_date = datetime.strptime(dates[0].rstrip(),"%Y-%m-%d %H:%M:%S")
                    end_date = datetime.strptime(dates[1].lstrip(),"%Y-%m-%d %H:%M:%S")
                except ValueError:
                    start_date = datetime.strptime(dates[0].rstrip(),"%Y-%m-%d %H:%M")
                    end_date = datetime.strptime(dates[1].lstrip(),"%Y-%m-%d %H:%M")
                
                #Get keys that inputed
                strategy_properties_keys = ['Initial capital', 'Order size', 'Pyramiding', 'Commission', 'Slippage', 'Verify price for limit orders', 'Margin for long positions', 'Margin for short positions','Recalculate After order is filled', 'Recalculate On every tick','Recalculate On bar close','Backtesting precision. Use bar magnifier']
                inputed = ['Trading range', 'Backtesting range','Symbol', 'Timeframe','Chart type','Point Value','Precision']
                strategy_properties = {}
                for key in strategy_properties_keys:
                    strategy_properties[key] = properties_data[key]
                concatenated = strategy_properties_keys + inputed
                #Get remained keys that not inputed
                strategy_inputs = {key: properties_data[key] for key in set(properties_data.keys()) - set(concatenated)}
                #make json
                json_prop = json.dumps(strategy_properties)
                json_int = json.dumps(strategy_inputs)

                #input to Properties
                inss = tbProperties.insert().values(**{
                "SymbolID": inputs,
                "TradingRange": properties_data['Trading range'],
                "BacktestingRange": properties_data['Backtesting range'],
                "Timeframe": properties_data['Timeframe'],
                "ChartType": properties_data['Chart type'],
                "PointValue": properties_data['Point Value'],
                "Precision": properties_data['Precision'],
                "BacktestStartYear": start_date.year,
                "BacktestStartDay": "{} - {}".format(start_date.day, start_date.month),
                "BacktestStopYear": end_date.year,
                "BacktestStopDay": "{} - {}".format(end_date.day, end_date.month),
                "StrategyInputs": json_int,
                "StrategyProperties": json_prop,

                }
                )   
                connection.execute(inss)
                connection.commit()
                connection.close()

            #insert StrategyName into tbStrategy
            with engine.connect() as connection:
                Strategy_Name = workbook.split(".")
                # Given string
                file_name = workbook

            # Define a regular expression pattern to match the date in the format YYYY-MM-DD
                date_pattern = r'\d{4}-\d{2}-\d{2}'

            # Use regular expression to search for the date pattern in the string
                match = re.search(date_pattern, file_name)

            # If a match is found, extract the date string
                if match:
                    date_str = match.group(0)
                    # Convert the date string to a datetime object
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                else:
                    print("No date in filename")
                strat_code = workbook.split("_")
                Strat_Insert = tbStrategy.insert().values(**{
                    "StrategyCode": strat_code[0],
                    "StrategyName": Strategy_Name[0],
                    "StrategyTestDate": date_obj,
                }
                )
                connection.execute(Strat_Insert)
                connection.commit()
                connection.close()

            #insert PropertiesID and StrategyID into tbSummary
            #get propID
            with engine.connect() as connection:
                latest_row_queryProperties = select(tbProperties.c.PropertiesID).order_by(tbProperties.c.PropertiesID.desc()).limit(1)
                prop = connection.execute(latest_row_queryProperties)
                property_ID = []
                for j in prop:
                    property_ID.append(j[0])
                
                connection.close()

            #get startID
            with engine.connect() as connection:
                latest_row_queryStrategy = select(tbStrategy.c.StrategyID).order_by(tbStrategy.c.StrategyID.desc()).limit(1)
                strat = connection.execute(latest_row_queryStrategy)
                strategy_ID = []
                for j in strat:
                    strategy_ID.append(j[0])
                connection.close()

            with engine.connect() as connection:
                insertProp = tbSummary.insert().values(**{
                    "PropertiesID": property_ID[0],
                    "StrategyID": strategy_ID[0],
                    "Currency": "USD",
                    "NetProfit": summary_data['Net Profit'],
                    "GrossProfit": summary_data['Gross Profit'],
                    "GrossLoss": summary_data['Gross Loss'],
                    "MaxRun_Up": summary_data['Max Run-up'],
                    "MaxDrawdown": summary_data['Max Drawdown'],
                    "BuyHoldReturn": summary_data['Buy & Hold Return'],
                    "SharpeRatio": summary_data['Sharpe Ratio'],
                    "SortinoRatio": summary_data['Sortino Ratio'],
                    "ProfitFactor": summary_data['Profit Factor'],
                    "MaxContractsHeld": summary_data['Max Contracts Held'],
                    "OpenPL": summary_data['Open PL'],
                    "CommissionPaid": summary_data['Commission Paid'],
                    "TotalClosedTrades": summary_data['Total Closed Trades'],
                    "TotalOpenTrades": summary_data['Total Open Trades'],
                    "NumberWinningTrades": summary_data['Number Winning Trades'],
                    "NumberLosingTrades": summary_data['Number Losing Trades'],
                    "PercentProfitable": summary_data['Percent Profitable'],
                    "AvgTrade": summary_data['Avg Trade'],
                    "AvgWinningTrade": summary_data['Avg Winning Trade'],
                    "AvgLosingTrade": summary_data['Avg Losing Trade'],
                    "RatioAvgWinAvgLoss": summary_data['Ratio Avg Win / Avg Loss'],
                    "LargestWinningTrade": summary_data['Largest Winning Trade'],
                    "LargestLosingTrade": summary_data['Largest Losing Trade'],
                    "AvgNumBarsinTrades": summary_data['Avg # Bars in Trades'],
                    "AvgNumBarsinWinningTrades": summary_data['Avg # Bars in Winning Trades'],
                    "AvgNumBarsinLosingTrades": summary_data['Avg # Bars in Losing Trades'],
                    "MarginCalls": summary_data['Margin Calls']
                })
                connection.execute(insertProp)
                connection.commit()
                connection.close()
                with engine.connect() as connection:
                    for i in list_data:
                        insertTradeList = tbTradeList.insert().values(**{
                            "StrategyID": strategy_ID[0],
                            "Type": i["Type"],
                            "Signal": i["Signal"],
                            "DateTime": i["Date/Time"],
                            "Price": i["Price USD"],
                            "Contracts": i["Contracts"],
                            "ProfitUSD": i["Profit USD"],
                            "ProfitPercent":i["Profit %"],
                            "CumProfitUSD": i["Cum. Profit USD"],
                            "CumProfitPercent": i["Cum. Profit %"],
                            "RunUpUSD": i["Run-up USD"],
                            "RunUpPercent": i["Run-up %"],
                            "DrawDownUSD": i["Drawdown USD"],
                            "DrawDownPercent": i["Drawdown %"],
                        })
                        connection.execute(insertTradeList)
                        connection.commit()
                    connection.close()
            try:    
                shutil.move(workbook_path, "/home/appcoder/excel_saved")
            except Exception:
                os.remove(workbook_path)
            graph_create(strategy_ID[0], put_in[1].strip())
    except Exception as e:
        with open("/home/appcoder/error-file/error.txt", "w") as file:
            file.write(str(e))






